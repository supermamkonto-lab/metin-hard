"""Modul eksportu danych do XLSX. Odpowiada za: - konwersje cen na liczby, - konwersje dat, - generowanie nazw plikow, - zapis danych do XLSX (z formatowaniem kolumn). """

from __future__ import annotations

import re
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from src.models import SaleEntry


def convert_date_to_iso(date_str: str) -> str:
    """Konwertuje '20:01 07.06.2026' na '20:01 2026-06-07'."""
    match = re.match(r"(\d{2}:\d{2})\s+(\d{2})\.(\d{2})\.(\d{4})", date_str)
    if match:
        return f"{match.group(1)} {match.group(4)}-{match.group(3)}-{match.group(2)}"
    return date_str



def format_price_display(price_str: str) -> str:
    """Formatuje cene do polskiego formatu tekstowego: 128.250,00

    Separator tysiecy (kropka) jest wpisany na stale w tekst, bo Excel
    renderuje separator grupowania liczb wg ustawien regionalnych
    systemu, na ktorym plik jest otwierany (nie da sie tego wymusic
    dla prawdziwie liczbowej komorki). Zapis jako tekst gwarantuje
    identyczny wyglad na kazdym komputerze.
    """
    cleaned = price_str.replace(" ", "").replace("\xa0", "")
    try:
        value = float(cleaned)
    except ValueError:
        try:
            value = float(cleaned.replace(",", "."))
        except ValueError:
            return price_str
    s = f"{value:,.2f}"
    s = s.replace(",", "\xa7").replace(".", ",").replace("\xa7", ".")
    return s




def export_xlsx(entries: list[SaleEntry], output_path: Path) -> None:
    """Zapisuje dane do pliku XLSX z formatowaniem."""
    headers = ["Postac", "Nazwa przedmiotu", "Ilosc", "Cena", "Typ ceny", "Data i godzina"]
    wb = Workbook()
    ws = wb.active
    ws.title = "Logi sprzedazy"

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2D5A27", end_color="2D5A27", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    for row_idx, entry in enumerate(entries, 2):
        values = [
            entry.character,
            entry.item_name,
            int(entry.quantity) if entry.quantity.isdigit() else entry.quantity,
            format_price_display(entry.price),
            entry.price_type,
            convert_date_to_iso(entry.date),
        ]
        for col, value in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.alignment = Alignment(vertical="center")
            cell.border = thin_border

    # Auto-fit kolumn
    for col_idx in range(1, len(headers) + 1):
        max_length = len(headers[col_idx - 1])
        for row_idx in range(2, len(entries) + 2):
            val = str(ws.cell(row=row_idx, column=col_idx).value or "")
            max_length = max(max_length, len(val))
        ws.column_dimensions[get_column_letter(col_idx)].width = max_length + 4

    for row_idx in range(2, len(entries) + 2):
        ws.cell(row=row_idx, column=3).alignment = Alignment(horizontal="center", vertical="center")
        cell_price = ws.cell(row=row_idx, column=4)
        cell_price.number_format = "@"
        cell_price.alignment = Alignment(horizontal="right", vertical="center")

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(entries) + 1}"
    wb.save(output_path)



def generate_filename(character_name: str) -> str:
    """Generuje nazwe pliku: POSTAC_DD_MM_RRRR_GG_MM (data i godzina systemowa)."""
    now = datetime.now()
    return f"{character_name}_{now.strftime('%d_%m_%Y_%H_%M')}"


def generate_combined_filename() -> str:
    """Generuje nazwe pliku zbiorczego (kilka postaci w jednym XLSX):
    WSZYSTKIE_POSTACIE_DD_MM_RRRR_GG_MM.
    """
    now = datetime.now()
    return f"WSZYSTKIE_POSTACIE_{now.strftime('%d_%m_%Y_%H_%M')}"
