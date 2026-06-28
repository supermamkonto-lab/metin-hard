"""METIN HARD LOG -- Eksporter Logow Sprzedazy v1.0

Profesjonalny interfejs uzytkownika.
Wyswietla kazdy etap dzialania programu.
"""

from __future__ import annotations

import csv
import re
import sys
import time
import traceback
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from playwright.sync_api import sync_playwright, TimeoutError as PlaywrightTimeoutError

from src.config import load_config
from src.auth import login
from src.characters import detect_characters
from src.scraper import scrape_sales_logs
from src.errors import ScraperError
from src.models import AppConfig, Character, SaleEntry


def print_header():
    print()
    print(" =========================================")
    print(" METIN HARD LOG")
    print(" Eksporter Logow Sprzedazy")
    print(" v1.0")
    print(" =========================================")
    print()


def print_step(step: int, total: int, message: str):
    print(f" [{step}/{total}] {message}")


def print_ok(detail: str = ""):
    if detail:
        print(f" {detail}")
    else:
        print(" OK")
    print()


def print_error_report(message: str, causes: list[str]):
    print()
    print(" =========================================")
    print(" BLAD")
    print(f" {message}")
    print()
    print(" Mozliwe przyczyny:")
    for cause in causes:
        print(f" * {cause}")
    print()
    # Zapis do logs/
    log_dir = Path("logs")
    log_dir.mkdir(exist_ok=True)
    log_file = log_dir / "error.log"
    with open(log_file, "a", encoding="utf-8") as lf:
        lf.write(chr(10) + chr(61)*50 + chr(10))
        lf.write(datetime.now().strftime('%Y-%m-%d %H:%M:%S') + chr(10))
        lf.write(message + chr(10))
        lf.write(traceback.format_exc())
    print(" Szczegoly zapisano do: logs\\error.log")
    print(" =========================================")


def convert_date_to_iso(date_str: str) -> str:
    """Konwertuje '20:01 07.06.2026' na '20:01 2026-06-07'."""
    match = re.match(r"(\d{2}:\d{2})\s+(\d{2})\.(\d{2})\.(\d{4})", date_str)
    if match:
        return f"{match.group(1)} {match.group(4)}-{match.group(3)}-{match.group(2)}"
    return date_str


def export_xlsx(entries: list[SaleEntry], output_path: Path) -> None:
    """Zapisuje dane do pliku XLSX z formatowaniem."""
    headers = ["Postac", "Nazwa przedmiotu", "Ilosc", "Cena", "Typ ceny", "Data i godzina"]
    wb = Workbook()
    ws = wb.active
    ws.title = "Logi sprzedazy"

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2D5A27", end_color="2D5A27", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    for row_idx, entry in enumerate(entries, 2):
        values = [
            entry.character,
            entry.item_name,
            int(entry.quantity) if entry.quantity.isdigit() else entry.quantity,
            int(entry.price) if entry.price.isdigit() else entry.price,
            entry.price_type,
            convert_date_to_iso(entry.date),
        ]
        for col, value in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.alignment = Alignment(vertical="center")
            cell.border = thin_border

    # Auto-fit kolumn
    for col_idx in range(1, len(headers) + 1):
        max_length = len(headers[col_idx - 1])
        for row_idx in range(2, len(entries) + 2):
            val = str(ws.cell(row=row_idx, column=col_idx).value or "")
            max_length = max(max_length, len(val))
        ws.column_dimensions[get_column_letter(col_idx)].width = max_length + 4

    for row_idx in range(2, len(entries) + 2):
        ws.cell(row=row_idx, column=4).number_format = '#,##0'
        ws.cell(row=row_idx, column=4).alignment = Alignment(horizontal="right", vertical="center")
        ws.cell(row=row_idx, column=3).alignment = Alignment(horizontal="center", vertical="center")

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(entries) + 1}"
    wb.save(output_path)


def export_csv_file(entries: list[SaleEntry], output_path: Path) -> None:
    """Zapisuje dane do pliku CSV z separatorem ;."""
    headers = ["Postac", "Nazwa przedmiotu", "Ilosc", "Cena", "Typ ceny", "Data i godzina"]
    with open(output_path, "w", encoding="utf-8-sig", newline="") as f:
        writer = csv.writer(f, delimiter=";")
        writer.writerow(headers)
        for entry in entries:
            writer.writerow([
                entry.character, entry.item_name, entry.quantity,
                entry.price, entry.price_type, convert_date_to_iso(entry.date),
            ])


def generate_filename(character_name: str) -> str:
    """Generuje nazwe pliku: POSTAC_DD_MM_RRRR_GG_MM."""
    now = datetime.now()
    return f"{character_name}_{now.strftime('%d_%m_%Y_%H_%M')}"


def main() -> int:
    """Glowna funkcja programu. Zwraca kod wyjscia (0=sukces, 1=blad)."""
    print_header()

    # [1/9] Konfiguracja
    print_step(1, 9, "Wczytywanie konfiguracji...")
    try:
        config = load_config(Path("config.toml"))
        print_ok()
    except ScraperError as e:
        print_error_report(
            "Nie udalo sie wczytac konfiguracji.",
            ["Brak pliku config.toml", "Niepoprawne dane w config.toml", "Brak wymaganego pola"]
        )
        return 1

    # [2/9] Uruchomienie przegladarki
    print_step(2, 9, "Uruchamianie przegladarki...")
    with sync_playwright() as p:
        browser = p.chromium.launch(headless=not config.show_browser)
        context = browser.new_context(
            user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
        )
        page = context.new_page()
        print_ok()

        try:
            # [3/9] Otwieranie strony
            print_step(3, 9, "Otwieranie strony logowania...")
            page.goto(config.ucp_url, timeout=config.timeout_ms)
            print_ok()

            # [4/9] Logowanie
            print_step(4, 9, "Logowanie do panelu UCP...")
            login(page, config)
            print_ok()

            # [5/9] Wykrywanie postaci
            print_step(5, 9, "Wyszukiwanie postaci...")
            characters = detect_characters(page)
            print_ok(f"Znaleziono {len(characters)} postaci:")
            for char in characters:
                print(f" * {char.name}")
            print()

            # [6/9] Pobieranie logow
            print_step(6, 9, "Pobieranie logow sprzedazy...")
            print()
            all_entries: list[SaleEntry] = []

            for char in characters:
                print(f" ------------------------------------------")
                print(f" Postac: {char.name}")
                print(f" ------------------------------------------")
                print(" Otwieranie logow...", end=" ")

                entries = scrape_sales_logs(page, char, config)

                if not entries:
                    print("Brak logow sprzedazy.")
                else:
                    print(f"{len(entries)} rekordow")
                    all_entries.extend(entries)
                print()

            # [7/9] Eksport XLSX
            print_step(7, 9, "Tworzenie pliku XLSX...")
            if all_entries:
                char_name = all_entries[0].character
                base_name = generate_filename(char_name)
                xlsx_path = config.output_dir / f"{base_name}.xlsx"
                export_xlsx(all_entries, xlsx_path)
                print_ok(xlsx_path.name)
            else:
                print_ok("Brak danych do eksportu.")

            # [8/9] Eksport CSV
            print_step(8, 9, "Tworzenie pliku CSV...")
            if all_entries:
                csv_path = config.output_dir / f"{base_name}.csv"
                export_csv_file(all_entries, csv_path)
                print_ok(csv_path.name)
            else:
                print_ok("Brak danych do eksportu.")

            # [9/9] Raport koncowy
            print_step(9, 9, "Generowanie raportu...")
            print()
            print(" =========================================")
            print(" Operacja zakonczona pomyslnie")
            print(" =========================================")
            print()
            print(f" Postaci: {len(characters)}")
            print(f" Lacznie rekordow: {len(all_entries)}")
            if all_entries:
                print()
                print(" Utworzono pliki:")
                print(f" + {xlsx_path.name}")
                print(f" + {csv_path.name}")
            print()
            print(f" Data eksportu: {datetime.now().strftime('%Y-%m-%d %H:%M')}")
            print(" Status: SUKCES")
            print(" =========================================")
            print()

            return 0

        except ScraperError as e:
            print_error_report(
                f"Blad na etapie: {e.stage}",
                ["Problem z polaczeniem", "Sesja wygasla", "Panel UCP zmienil strukture", "Sprawdz diagnostyke"]
            )
            return 1

        except Exception as e:
            print_error_report(
                "Nieoczekiwany blad programu.",
                ["Sprawdz polaczenie internetowe", "Uruchom diagnostyke", "Skontaktuj sie z administratorem"]
            )
            return 1

        finally:
            browser.close()


if __name__ == "__main__":
    sys.exit(main())
