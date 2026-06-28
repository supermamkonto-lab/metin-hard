"""Metin HARD Log Extractor — główny punkt wejścia.

Orkiestruje cały proces:
1. Wczytanie konfiguracji
2. Logowanie do UCP
3. Wykrycie postaci
4. Pobranie logów sprzedaży
5. Eksport do XLSX i CSV
6. Raport zakończenia
"""

from __future__ import annotations

import csv
import re
import sys
import time
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from playwright.sync_api import sync_playwright

from src.config import load_config
from src.auth import login
from src.characters import detect_characters
from src.scraper import scrape_sales_logs
from src.errors import ScraperError
from src.models import AppConfig, Character, SaleEntry


def convert_date_to_iso(date_str: str) -> str:
    """Konwertuje '20:01 07.06.2026' na '20:01 2026-06-07'."""
    match = re.match(r"(\d{2}:\d{2})\s+(\d{2})\.(\d{2})\.(\d{4})", date_str)
    if match:
        return f"{match.group(1)} {match.group(4)}-{match.group(3)}-{match.group(2)}"
    return date_str


def export_xlsx(entries: list[SaleEntry], output_path: Path) -> None:
    """Zapisuje dane do pliku XLSX z formatowaniem."""
    headers = ["Postać", "Nazwa przedmiotu", "Ilość", "Cena", "Typ ceny", "Data i godzina"]

    wb = Workbook()
    ws = wb.active
    ws.title = "Logi sprzedaży"

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2D5A27", end_color="2D5A27", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    for row_idx, entry in enumerate(entries, 2):
        date_converted = convert_date_to_iso(entry.date)
        values = [
            entry.character,
            entry.item_name,
            int(entry.quantity) if entry.quantity.isdigit() else entry.quantity,
            int(entry.price) if entry.price.isdigit() else entry.price,
            entry.price_type,
            date_converted,
        ]
        for col, value in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.alignment = Alignment(vertical="center")
            cell.border = thin_border

    # Auto-fit kolumn
    for col_idx in range(1, len(headers) + 1):
        max_length = len(headers[col_idx - 1])
        for row_idx in range(2, len(entries) + 2):
            val = str(ws.cell(row=row_idx, column=col_idx).value or "")
            max_length = max(max_length, len(val))
        ws.column_dimensions[get_column_letter(col_idx)].width = max_length + 4

    # Cena — separator tysięcy
    for row_idx in range(2, len(entries) + 2):
        ws.cell(row=row_idx, column=4).number_format = '#,##0'
        ws.cell(row=row_idx, column=4).alignment = Alignment(horizontal="right", vertical="center")

    # Ilość — wyśrodkowana
    for row_idx in range(2, len(entries) + 2):
        ws.cell(row=row_idx, column=3).alignment = Alignment(horizontal="center", vertical="center")

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(entries) + 1}"

    wb.save(output_path)


def export_csv(entries: list[SaleEntry], output_path: Path) -> None:
    """Zapisuje dane do pliku CSV z separatorem ;."""
    headers = ["Postać", "Nazwa przedmiotu", "Ilość", "Cena", "Typ ceny", "Data i godzina"]

    with open(output_path, "w", encoding="utf-8-sig", newline="") as f:
        writer = csv.writer(f, delimiter=";")
        writer.writerow(headers)
        for entry in entries:
            date_converted = convert_date_to_iso(entry.date)
            writer.writerow([
                entry.character, entry.item_name, entry.quantity,
                entry.price, entry.price_type, date_converted,
            ])


def generate_filename(character_name: str) -> str:
    """Generuje nazwę pliku: POSTAĆ_DD_MM_RRRR_GG_MM."""
    now = datetime.now()
    return f"{character_name}_{now.strftime('%d_%m_%Y_%H_%M')}"


def main() -> int:
    """Główna funkcja programu. Zwraca kod wyjścia (0=sukces, 1=błąd)."""
    print()
    print("  ==========================================")
    print("  METIN HARD Log Extractor")
    print("  ==========================================")
    print()

    # 1. Konfiguracja
    try:
        config = load_config(Path("config.toml"))
    except ScraperError as e:
        print(f"  BLAD: {e.message}")
        return 1

    # 2. Uruchomienie przeglądarki i logowanie
    with sync_playwright() as p:
        browser = p.chromium.launch(headless=False)
        context = browser.new_context(
            user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
        )
        page = context.new_page()

        try:
            # Logowanie
            print("  [1/4] Logowanie do panelu UCP...")
            login(page, config)
            print("        OK")

            # Wykrywanie postaci
            print("  [2/4] Wykrywanie postaci...")
            characters = detect_characters(page)
            print(f"        Znaleziono: {len(characters)} postaci")

            # Pobieranie logów
            print("  [3/4] Pobieranie logow sprzedazy...")
            all_entries: list[SaleEntry] = []

            for char in characters:
                print(f"        - {char.name}...", end=" ")
                entries = scrape_sales_logs(page, char, config)
                all_entries.extend(entries)
                print(f"{len(entries)} wpisow")

            # Eksport
            print("  [4/4] Eksport do plikow...")

            if not all_entries:
                print("        Brak danych do eksportu.")
                print()
                print("  ==========================================")
                print("  Status: BRAK DANYCH")
                print("  ==========================================")
                return 0

            # Nazwa na podstawie pierwszej postaci z danymi
            char_with_data = all_entries[0].character
            base_name = generate_filename(char_with_data)

            xlsx_path = config.output_dir / f"{base_name}.xlsx"
            csv_path = config.output_dir / f"{base_name}.csv"

            export_xlsx(all_entries, xlsx_path)
            export_csv(all_entries, csv_path)

            print(f"        {xlsx_path.name}")
            print(f"        {csv_path.name}")

            # Raport
            print()
            print("  ==========================================")
            print("  Metin HARD Log Extractor — Eksport zakonczony")
            print("  ==========================================")
            print()
            print(f"  Postaci:          {len(characters)}")
            print(f"  Rekordow:         {len(all_entries)}")
            print(f"  Data eksportu:    {datetime.now().strftime('%H:%M %Y-%m-%d')}")
            print()
            print(f"  Pliki:")
            print(f"    {xlsx_path.name}  (Excel)")
            print(f"    {csv_path.name}   (CSV)")
            print()
            print(f"  Status: SUCCESS")
            print("  ==========================================")

            return 0

        except ScraperError as e:
            print()
            print(f"  BLAD [{e.stage}]: {e.message}")
            print()
            print("  ==========================================")
            print("  Status: BLAD")
            print("  ==========================================")
            return 1

        except Exception as e:
            print()
            print(f"  NIEOCZEKIWANY BLAD: {e}")
            print()
            print("  ==========================================")
            print("  Status: BLAD")
            print("  ==========================================")
            return 1

        finally:
            browser.close()


if __name__ == "__main__":
    sys.exit(main())
